#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Batch rename image files using an Excel mapping file.

Default usage:
    python rename_images.py

Preview only, no renaming:
    python rename_images.py --excel mapping.xlsx --image-dir image

Actually rename files:
    python rename_images.py --excel mapping.xlsx --image-dir image --apply

Excel format:
    Column A = Chinese name
    Column B = English file name
    Blank rows are OK.
    A header row is OK.
"""

from __future__ import annotations

from pathlib import Path
import argparse
import csv
import re
import sys
from typing import Dict, List, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("缺少套件 openpyxl。請先執行：python -m pip install openpyxl", file=sys.stderr)
    raise


IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
WINDOWS_FORBIDDEN_CHARS = r'[\\/:*?"<>|]'


def safe_filename(name: str, use_underscores: bool = False) -> str:
    """Convert a human-readable English name into a Windows-safe filename stem."""
    name = str(name or "").strip()
    name = re.sub(r"\s+", " ", name)
    name = re.sub(WINDOWS_FORBIDDEN_CHARS, "", name)
    name = name.strip(" .")
    if use_underscores:
        name = name.replace(" ", "_")
    return name


def strip_number_prefix(filename_stem: str) -> str:
    """01_報紙 -> 報紙; 03-剩菜剩飯 -> 剩菜剩飯; 08 陶瓷碎片 -> 陶瓷碎片"""
    return re.sub(r"^\d+[_\-\s]*", "", filename_stem).strip()


def normalize_chinese_name(name: str) -> str:
    """
    Normalize Chinese labels for matching.

    Examples:
        堅果殼 (如核桃殼) -> 堅果殼
        堅果殼（如核桃殼） -> 堅果殼
        骨頭渣 (小碎骨) -> 骨頭渣
    """
    name = str(name or "").strip()
    name = re.sub(r"（.*?）", "", name)
    name = re.sub(r"\(.*?\)", "", name)
    name = re.sub(r"\[.*?\]", "", name)
    name = re.sub(r"【.*?】", "", name)
    name = re.sub(r"\s+", "", name)
    return name.strip()


def is_header_row(chinese: str, english: str) -> bool:
    c = str(chinese or "").strip().lower()
    e = str(english or "").strip().lower()
    return (
        c in {"中文", "中文名", "chinese", "chinese name", "原名"}
        or e in {"英文", "英文名", "英文檔名", "english", "english name", "filename", "file name"}
    )


def make_unique_path(path: Path) -> Path:
    """Avoid overwriting an existing file by appending _2, _3, etc."""
    if not path.exists():
        return path

    folder = path.parent
    stem = path.stem
    suffix = path.suffix
    counter = 2
    while True:
        candidate = folder / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def write_csv(path: Path, rows: List[dict], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def read_mapping(excel_path: Path, sheet_name: str | None, use_underscores: bool):
    if not excel_path.exists():
        raise FileNotFoundError(f"找不到 Excel 檔案：{excel_path}")

    wb = load_workbook(excel_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Excel 裡找不到工作表：{sheet_name}。可用工作表：{', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    mapping: Dict[str, dict] = {}
    incomplete_rows: List[dict] = []
    duplicate_rows: List[dict] = []
    valid_rows = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=1, max_col=2, values_only=True), start=1):
        chinese_raw, english_raw = row[0], row[1]
        chinese = "" if chinese_raw is None else str(chinese_raw).strip()
        english = "" if english_raw is None else str(english_raw).strip()

        if not chinese and not english:
            continue

        if is_header_row(chinese, english):
            continue

        if bool(chinese) ^ bool(english):
            incomplete_rows.append({
                "Excel列號": row_num,
                "中文名": chinese,
                "英文檔名": english,
                "原因": "這列只填了一邊，已跳過",
            })
            continue

        normalized = normalize_chinese_name(chinese)
        safe_english = safe_filename(english, use_underscores=use_underscores)

        if not normalized or not safe_english:
            incomplete_rows.append({
                "Excel列號": row_num,
                "中文名": chinese,
                "英文檔名": english,
                "原因": "中文名或英文檔名清理後變成空白，已跳過",
            })
            continue

        valid_rows += 1
        if normalized in mapping:
            duplicate_rows.append({
                "Excel列號": row_num,
                "中文名": chinese,
                "標準化中文名": normalized,
                "英文檔名": safe_english,
                "原因": "標準化後中文名重複，已跳過後面這筆",
            })
            continue

        mapping[normalized] = {
            "excel_row": row_num,
            "original_chinese": chinese,
            "english_name": safe_english,
        }

    return mapping, incomplete_rows, duplicate_rows, valid_rows, ws.title


def collect_images(image_dir: Path) -> List[Path]:
    if not image_dir.exists():
        raise FileNotFoundError(f"找不到圖片資料夾：{image_dir}")
    if not image_dir.is_dir():
        raise NotADirectoryError(f"這不是資料夾：{image_dir}")

    return sorted(
        [p for p in image_dir.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS],
        key=lambda p: p.name.lower(),
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="用 Excel 對照表批次改圖片英文檔名")
    parser.add_argument("--excel", default="mapping.xlsx", help="Excel 對照表路徑，預設 mapping.xlsx")
    parser.add_argument("--sheet", default=None, help="工作表名稱；不填就用第一個工作表")
    parser.add_argument("--image-dir", default="image", help="圖片資料夾，預設 image")
    parser.add_argument("--report-dir", default="rename_reports", help="報告輸出資料夾，預設 rename_reports")
    parser.add_argument("--apply", action="store_true", help="真的改名；不加這個參數時只會預覽")
    parser.add_argument("--use-underscores", action="store_true", help="把英文檔名空格改成底線")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    image_dir = Path(args.image_dir)
    report_dir = Path(args.report_dir)

    mapping, incomplete_rows, duplicate_rows, valid_rows, used_sheet = read_mapping(
        excel_path, args.sheet, args.use_underscores
    )
    image_files = collect_images(image_dir)

    renamed_report: List[dict] = []
    unmatched_images: List[dict] = []
    found_chinese_names = set()

    for file in image_files:
        original_name = file.name
        chinese_name = strip_number_prefix(file.stem)
        normalized_chinese = normalize_chinese_name(chinese_name)
        found_chinese_names.add(normalized_chinese)

        if normalized_chinese not in mapping:
            unmatched_images.append({
                "原圖片檔名": original_name,
                "識別出的中文名": chinese_name,
                "標準化中文名": normalized_chinese,
                "原因": "Excel 找不到對應中文名",
            })
            continue

        english_name = mapping[normalized_chinese]["english_name"]
        target_path = file.with_name(english_name + file.suffix.lower())
        target_path = make_unique_path(target_path)

        status = "預覽，尚未改名"
        error = ""
        if args.apply:
            try:
                file.rename(target_path)
                status = "已改名"
            except Exception as exc:
                status = "改名失敗"
                error = str(exc)

        renamed_report.append({
            "原圖片檔名": original_name,
            "識別出的中文名": chinese_name,
            "Excel中文名": mapping[normalized_chinese]["original_chinese"],
            "Excel列號": mapping[normalized_chinese]["excel_row"],
            "新圖片檔名": target_path.name,
            "狀態": status,
            "錯誤訊息": error,
        })

    missing_images: List[dict] = []
    for normalized_chinese, data in mapping.items():
        if normalized_chinese not in found_chinese_names:
            missing_images.append({
                "Excel中文名": data["original_chinese"],
                "標準化中文名": normalized_chinese,
                "英文檔名": data["english_name"],
                "Excel列號": data["excel_row"],
                "原因": "Excel 有這個項目，但圖片資料夾沒有對應圖片",
            })

    write_csv(report_dir / "renamed_report.csv", renamed_report, [
        "原圖片檔名", "識別出的中文名", "Excel中文名", "Excel列號", "新圖片檔名", "狀態", "錯誤訊息"
    ])
    write_csv(report_dir / "unmatched_images.csv", unmatched_images, [
        "原圖片檔名", "識別出的中文名", "標準化中文名", "原因"
    ])
    write_csv(report_dir / "missing_images.csv", missing_images, [
        "Excel中文名", "標準化中文名", "英文檔名", "Excel列號", "原因"
    ])

    if incomplete_rows:
        write_csv(report_dir / "incomplete_excel_rows.csv", incomplete_rows, [
            "Excel列號", "中文名", "英文檔名", "原因"
        ])
    if duplicate_rows:
        write_csv(report_dir / "duplicate_chinese_rows.csv", duplicate_rows, [
            "Excel列號", "中文名", "標準化中文名", "英文檔名", "原因"
        ])

    summary = [
        "=========================",
        "處理完成",
        "=========================",
        f"模式：{'實際改名' if args.apply else '預覽，不會改名'}",
        f"Excel：{excel_path}",
        f"工作表：{used_sheet}",
        f"圖片資料夾：{image_dir}",
        f"報告資料夾：{report_dir}",
        "",
        f"Excel有效對照列數：{len(mapping)}",
        f"Excel有效列數含重複：{valid_rows}",
        f"圖片總數：{len(image_files)}",
        f"可改名圖片數：{len(renamed_report)}",
        f"圖片找不到 Excel 對應數：{len(unmatched_images)}",
        f"Excel 有但圖片沒有數：{len(missing_images)}",
        f"Excel 只填一邊列數：{len(incomplete_rows)}",
        f"重複中文列數：{len(duplicate_rows)}",
        "",
        "請查看 CSV 報告：",
        "- renamed_report.csv",
        "- unmatched_images.csv",
        "- missing_images.csv",
        "- incomplete_excel_rows.csv（如果有）",
        "- duplicate_chinese_rows.csv（如果有）",
    ]
    (report_dir / "summary.txt").write_text("\n".join(summary), encoding="utf-8-sig")

    print("\n".join(summary))

    if not args.apply:
        print("\n目前只是預覽。確認 rename_reports/renamed_report.csv 沒問題後，再執行：")
        print(f"python rename_images.py --excel {excel_path} --image-dir {image_dir} --apply")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
